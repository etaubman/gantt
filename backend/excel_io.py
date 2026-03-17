"""Excel export and import for full project state."""
import io
import json
import os
import tempfile
import uuid
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from backend.database import get_conn, DB_PATH

SCHEMA_VERSION = "3"
APP_VERSION = "1.0.0"


def _as_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in ("1", "true", "yes", "y")


def _build_task_tree(tasks: list[dict]) -> list[dict]:
    by_parent: dict[str | None, list[dict]] = defaultdict(list)
    for task in tasks:
        by_parent[task.get("parent_task_uid")].append(task)
    for children in by_parent.values():
        children.sort(key=lambda item: (item.get("sort_order", 0), item.get("created_at") or ""))
    ordered: list[dict] = []

    def walk(parent_uid: str | None, depth: int, prefix: str) -> None:
        children = by_parent.get(parent_uid, [])
        for index, task in enumerate(children, start=1):
            hierarchy_number = f"{prefix}.{index}" if prefix else str(index)
            task_copy = dict(task)
            task_copy["depth"] = depth
            task_copy["hierarchy_number"] = hierarchy_number
            ordered.append(task_copy)
            walk(task["uid"], depth + 1, hierarchy_number)

    walk(None, 0, "")
    return ordered


def _apply_sheet_chrome(ws, widths: dict[str, int] | None = None, freeze_cell: str = "A2") -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F3552")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = True
    if widths:
        for column_letter, width in widths.items():
            ws.column_dimensions[column_letter].width = width


def _wrap_columns(ws, column_letters: list[str], start_row: int = 2) -> None:
    for column_letter in column_letters:
        for row in range(start_row, ws.max_row + 1):
            ws[f"{column_letter}{row}"].alignment = Alignment(vertical="top", wrap_text=True)


def _status_fill(status: str | None) -> PatternFill | None:
    tones = {
        "complete": "D8F3DC",
        "in_progress": "DBEAFE",
        "not_started": "E5E7EB",
        "blocked": "FEE2E2",
        "cancelled": "F3E8FF",
        "open": "FEE2E2",
        "mitigated": "FEF3C7",
        "closed": "D8F3DC",
    }
    color = tones.get((status or "").strip().lower())
    if not color:
        return None
    return PatternFill(fill_type="solid", fgColor=color)


def export_project_to_xlsx(project_uid: str) -> str | None:
    """Export one project (and its data) to an xlsx file. Returns file path or None if project not found."""
    exported_at = datetime.utcnow().isoformat() + "Z"
    exported_at_safe = exported_at.replace(":", "-")
    with get_conn() as conn:
        proj = conn.execute("SELECT uid, name, created_at FROM projects WHERE uid = ?", (project_uid,)).fetchone()
        if not proj:
            return None
        projects = [dict(proj)]
        tasks = [dict(r) for r in conn.execute(
            """SELECT uid, project_uid, parent_task_uid, name, description, accountable_person, responsible_party,
                      start_date, end_date, is_milestone, status, progress, sort_order, scheduling_mode,
                      is_deleted, deleted_at, deleted_by, created_at, updated_at
               FROM tasks WHERE project_uid = ?""",
            (project_uid,),
        ).fetchall()]
        task_uids = {t["uid"] for t in tasks}
        deps = [dict(r) for r in conn.execute(
            "SELECT uid, project_uid, predecessor_task_uid, successor_task_uid, dependency_type, created_at FROM dependencies WHERE project_uid = ?",
            (project_uid,),
        ).fetchall()]
        rag = []
        comments = []
        risks = []
        audit_events = [dict(r) for r in conn.execute(
            """SELECT uid, actor_employee_id, action_type, entity_type, entity_uid, task_uid, task_name,
                      prior_value, new_value, metadata, created_at
               FROM audit_events ORDER BY created_at ASC"""
        ).fetchall()]
        edit_lock = conn.execute(
            "SELECT lock_name, employee_id, locked_at, updated_at FROM edit_lock WHERE lock_name = ?",
            ("workspace",),
        ).fetchone()
        for t in tasks:
            rag.extend([dict(r) for r in conn.execute("SELECT uid, task_uid, status, rationale, path_to_green, created_at FROM rag_statuses WHERE task_uid = ?", (t["uid"],)).fetchall()])
            comments.extend([dict(r) for r in conn.execute("SELECT uid, task_uid, author, comment_text, created_at FROM comments WHERE task_uid = ?", (t["uid"],)).fetchall()])
            risks.extend([dict(r) for r in conn.execute("SELECT uid, task_uid, title, description, severity, status, owner, mitigation_plan, created_at, updated_at FROM risks WHERE task_uid = ?", (t["uid"],)).fetchall()])

    wb = Workbook()
    # Metadata
    ws_meta = wb.active
    ws_meta.title = "Metadata"
    ws_meta.append(["schema_version", "exported_at", "application_version"])
    ws_meta.append([SCHEMA_VERSION, exported_at, APP_VERSION])

    # Projects
    ws_proj = wb.create_sheet("Projects")
    ws_proj.append(["Project UID", "Project Name", "Created At"])
    for p in projects:
        ws_proj.append([p["uid"], p["name"], p["created_at"]])

    # Tasks
    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks.append(["Task UID", "Project UID", "Parent Task UID", "Name", "Description", "Accountable Person", "Responsible Party", "Start Date", "End Date", "Is Milestone", "Status", "Progress", "Sort Order", "Scheduling Mode", "Is Deleted", "Deleted At", "Deleted By", "Created At", "Updated At"])
    for t in tasks:
        ws_tasks.append([t["uid"], t["project_uid"], t["parent_task_uid"] or "", t["name"], t["description"] or "", t["accountable_person"] or "", t["responsible_party"] or "", t["start_date"] or "", t["end_date"] or "", bool(t.get("is_milestone")), t["status"], t["progress"], t["sort_order"], t.get("scheduling_mode") or "fixed", bool(t.get("is_deleted")), t.get("deleted_at") or "", t.get("deleted_by") or "", t["created_at"], t["updated_at"]])

    # Dependencies
    ws_dep = wb.create_sheet("Dependencies")
    ws_dep.append(["Dependency UID", "Project UID", "Predecessor Task UID", "Successor Task UID", "Dependency Type", "Created At"])
    for d in deps:
        ws_dep.append([d["uid"], d["project_uid"], d["predecessor_task_uid"], d["successor_task_uid"], d["dependency_type"], d["created_at"]])

    # RAG
    ws_rag = wb.create_sheet("RAG Status History")
    ws_rag.append(["RAG UID", "Task UID", "Status", "Rationale", "Path To Green", "Created At"])
    for r in rag:
        ws_rag.append([r["uid"], r["task_uid"], r["status"], r["rationale"] or "", r.get("path_to_green") or "", r["created_at"]])

    # Comments
    ws_com = wb.create_sheet("Comments")
    ws_com.append(["Comment UID", "Task UID", "Author", "Comment Text", "Created At"])
    for c in comments:
        ws_com.append([c["uid"], c["task_uid"], c["author"], c["comment_text"] or "", c["created_at"]])

    # Risks
    ws_risk = wb.create_sheet("Risks")
    ws_risk.append(["Risk UID", "Task UID", "Title", "Description", "Severity", "Status", "Owner", "Mitigation Plan", "Created At", "Updated At"])
    for r in risks:
        ws_risk.append([r["uid"], r["task_uid"], r["title"], r["description"] or "", r["severity"], r["status"], r["owner"] or "", r["mitigation_plan"] or "", r["created_at"], r["updated_at"]])

    ws_audit = wb.create_sheet("Audit Log")
    ws_audit.append(["Audit UID", "Actor Employee ID", "Action Type", "Entity Type", "Entity UID", "Task UID", "Task Name", "Prior Value", "New Value", "Metadata", "Created At"])
    for event in audit_events:
        ws_audit.append([
            event["uid"],
            event["actor_employee_id"],
            event["action_type"],
            event["entity_type"],
            event.get("entity_uid") or "",
            event.get("task_uid") or "",
            event.get("task_name") or "",
            event.get("prior_value") or "",
            event.get("new_value") or "",
            event.get("metadata") or "",
            event["created_at"],
        ])

    ws_lock = wb.create_sheet("Edit Lock")
    ws_lock.append(["Lock Name", "Employee ID", "Locked At", "Updated At"])
    if edit_lock:
        ws_lock.append([
            edit_lock["lock_name"],
            edit_lock["employee_id"],
            edit_lock["locked_at"],
            edit_lock["updated_at"],
        ])

    dir_path = os.path.dirname(DB_PATH) or tempfile.gettempdir()
    os.makedirs(dir_path, exist_ok=True)
    path = os.path.join(dir_path, f"export_{project_uid[:8]}_{exported_at_safe}.xlsx")
    wb.save(path)
    return path


def export_project_report_to_xlsx(project_uid: str) -> str | None:
    """Export one project to a human-readable offline report workbook."""
    exported_at = datetime.utcnow().isoformat() + "Z"
    exported_at_safe = exported_at.replace(":", "-")
    with get_conn() as conn:
        proj = conn.execute("SELECT uid, name, created_at FROM projects WHERE uid = ?", (project_uid,)).fetchone()
        if not proj:
            return None
        project = dict(proj)
        tasks = [dict(r) for r in conn.execute(
            """SELECT uid, project_uid, parent_task_uid, name, description, accountable_person, responsible_party,
                      start_date, end_date, is_milestone, status, progress, sort_order, scheduling_mode,
                      is_deleted, deleted_at, deleted_by, created_at, updated_at
               FROM tasks WHERE project_uid = ?
               ORDER BY sort_order ASC, created_at ASC""",
            (project_uid,),
        ).fetchall()]
        deps = [dict(r) for r in conn.execute(
            "SELECT uid, predecessor_task_uid, successor_task_uid, dependency_type, created_at FROM dependencies WHERE project_uid = ? ORDER BY created_at ASC",
            (project_uid,),
        ).fetchall()]
        comments = [dict(r) for r in conn.execute(
            """SELECT uid, task_uid, author, comment_text, created_at
               FROM comments
               WHERE task_uid IN (SELECT uid FROM tasks WHERE project_uid = ?)
               ORDER BY created_at ASC""",
            (project_uid,),
        ).fetchall()]
        risks = [dict(r) for r in conn.execute(
            """SELECT uid, task_uid, title, description, severity, status, owner, mitigation_plan, created_at, updated_at
               FROM risks
               WHERE task_uid IN (SELECT uid FROM tasks WHERE project_uid = ?)
               ORDER BY created_at ASC""",
            (project_uid,),
        ).fetchall()]
        rag_rows = [dict(r) for r in conn.execute(
            """SELECT uid, task_uid, status, rationale, path_to_green, created_at
               FROM rag_statuses
               WHERE task_uid IN (SELECT uid FROM tasks WHERE project_uid = ?)
               ORDER BY created_at ASC""",
            (project_uid,),
        ).fetchall()]

    task_tree = _build_task_tree(tasks)
    task_by_uid = {task["uid"]: task for task in tasks}
    latest_comment_by_task: dict[str, dict] = {}
    for comment in comments:
        latest_comment_by_task[comment["task_uid"]] = comment
    latest_rag_by_task: dict[str, dict] = {}
    for rag in rag_rows:
        latest_rag_by_task[rag["task_uid"]] = rag
    open_risks_by_task: dict[str, list[dict]] = defaultdict(list)
    for risk in risks:
        if (risk.get("status") or "").lower() != "closed":
            open_risks_by_task[risk["task_uid"]].append(risk)
    predecessor_names_by_task: dict[str, list[str]] = defaultdict(list)
    successor_names_by_task: dict[str, list[str]] = defaultdict(list)
    for dep in deps:
        predecessor = task_by_uid.get(dep["predecessor_task_uid"])
        successor = task_by_uid.get(dep["successor_task_uid"])
        if predecessor and successor:
            predecessor_names_by_task[successor["uid"]].append(predecessor["name"])
            successor_names_by_task[predecessor["uid"]].append(successor["name"])

    wb = Workbook()

    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.append(["Project", project["name"]])
    ws_overview.append(["Project UID", project["uid"]])
    ws_overview.append(["Created At", project["created_at"]])
    ws_overview.append(["Exported At", exported_at])
    ws_overview.append(["Visible Tasks", sum(1 for task in tasks if not _as_bool(task.get("is_deleted")))])
    ws_overview.append(["Open Risks", sum(len(items) for items in open_risks_by_task.values())])
    ws_overview.append(["Latest Comments", len(latest_comment_by_task)])
    for cell in ws_overview["A"]:
        cell.font = Font(bold=True)
    ws_overview.column_dimensions["A"].width = 20
    ws_overview.column_dimensions["B"].width = 90
    _wrap_columns(ws_overview, ["B"], start_row=1)

    ws_tasks = wb.create_sheet("Task Report")
    ws_tasks.append([
        "Hierarchy",
        "Task",
        "Task UID",
        "Parent Task UID",
        "Status",
        "Progress %",
        "RAG",
        "Milestone",
        "Start",
        "End",
        "Accountable",
        "Responsible",
        "Latest Comment",
        "Latest Comment Author",
        "Latest Comment At",
        "Open Risks",
        "Predecessors",
        "Successors",
        "Description",
    ])
    for task in task_tree:
        latest_comment = latest_comment_by_task.get(task["uid"])
        latest_rag = latest_rag_by_task.get(task["uid"])
        open_risks = open_risks_by_task.get(task["uid"], [])
        risk_summary = "\n".join(
            f"{risk['title']} ({risk['severity']}, {risk['status']})"
            for risk in open_risks
        )
        ws_tasks.append([
            task.get("hierarchy_number") or "",
            task["name"],
            task["uid"],
            task.get("parent_task_uid") or "",
            task.get("status") or "",
            task.get("progress") if task.get("progress") is not None else 0,
            (latest_rag or {}).get("status") or "",
            "Yes" if _as_bool(task.get("is_milestone")) else "",
            task.get("start_date") or "",
            task.get("end_date") or "",
            task.get("accountable_person") or "",
            task.get("responsible_party") or "",
            (latest_comment or {}).get("comment_text") or "",
            (latest_comment or {}).get("author") or "",
            (latest_comment or {}).get("created_at") or "",
            risk_summary,
            "\n".join(predecessor_names_by_task.get(task["uid"], [])),
            "\n".join(successor_names_by_task.get(task["uid"], [])),
            task.get("description") or "",
        ])
    _apply_sheet_chrome(
        ws_tasks,
        {
            "A": 12, "B": 32, "C": 38, "D": 38, "E": 14, "F": 12, "G": 12, "H": 10,
            "I": 12, "J": 12, "K": 18, "L": 18, "M": 48, "N": 18, "O": 20, "P": 42,
            "Q": 28, "R": 28, "S": 52,
        },
    )
    _wrap_columns(ws_tasks, ["M", "P", "Q", "R", "S"])
    for row_idx in range(2, ws_tasks.max_row + 1):
        depth = 0
        hierarchy = ws_tasks[f"A{row_idx}"].value or ""
        if hierarchy:
            depth = str(hierarchy).count(".")
        task_cell = ws_tasks[f"B{row_idx}"]
        task_cell.alignment = Alignment(vertical="top", wrap_text=True, indent=min(depth, 8))
        status_fill = _status_fill(ws_tasks[f"E{row_idx}"].value)
        if status_fill:
            ws_tasks[f"E{row_idx}"].fill = status_fill

    ws_risks = wb.create_sheet("Open Risks")
    ws_risks.append([
        "Task",
        "Task UID",
        "Risk",
        "Severity",
        "Status",
        "Owner",
        "Mitigation Plan",
        "Description",
        "Created At",
        "Updated At",
    ])
    for task in task_tree:
        for risk in open_risks_by_task.get(task["uid"], []):
            ws_risks.append([
                task["name"],
                task["uid"],
                risk.get("title") or "",
                risk.get("severity") or "",
                risk.get("status") or "",
                risk.get("owner") or "",
                risk.get("mitigation_plan") or "",
                risk.get("description") or "",
                risk.get("created_at") or "",
                risk.get("updated_at") or "",
            ])
    if ws_risks.max_row == 1:
        ws_risks.append(["No open risks", "", "", "", "", "", "", "", "", ""])
    _apply_sheet_chrome(
        ws_risks,
        {"A": 28, "B": 38, "C": 30, "D": 12, "E": 12, "F": 18, "G": 42, "H": 42, "I": 20, "J": 20},
    )
    _wrap_columns(ws_risks, ["G", "H"])
    for row_idx in range(2, ws_risks.max_row + 1):
        status_fill = _status_fill(ws_risks[f"E{row_idx}"].value)
        if status_fill:
            ws_risks[f"E{row_idx}"].fill = status_fill

    ws_comments = wb.create_sheet("Latest Comments")
    ws_comments.append(["Task", "Task UID", "Author", "Comment", "Comment At"])
    for task in task_tree:
        latest_comment = latest_comment_by_task.get(task["uid"])
        if not latest_comment:
            continue
        ws_comments.append([
            task["name"],
            task["uid"],
            latest_comment.get("author") or "",
            latest_comment.get("comment_text") or "",
            latest_comment.get("created_at") or "",
        ])
    if ws_comments.max_row == 1:
        ws_comments.append(["No comments yet", "", "", "", ""])
    _apply_sheet_chrome(ws_comments, {"A": 28, "B": 38, "C": 18, "D": 70, "E": 20})
    _wrap_columns(ws_comments, ["D"])

    dir_path = os.path.dirname(DB_PATH) or tempfile.gettempdir()
    os.makedirs(dir_path, exist_ok=True)
    path = os.path.join(dir_path, f"report_{project_uid[:8]}_{exported_at_safe}.xlsx")
    wb.save(path)
    return path


def import_xlsx(content: bytes) -> dict:
    """Import from workbook bytes. Creates new projects (by UID). Returns summary."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = set(wb.sheetnames)

    def get_sheet(name: str):
        if name not in sheet_names:
            raise ValueError(f"Missing sheet: {name}")
        return wb[name]

    def get_optional_sheet(name: str):
        return wb[name] if name in sheet_names else None

    # Metadata optional
    # Projects
    ws_proj = get_sheet("Projects")
    rows_proj = list(ws_proj.iter_rows(min_row=2, values_only=True))
    projects = []
    for row in rows_proj:
        if row and row[0]:
            projects.append({"uid": str(row[0]).strip(), "name": str(row[1]) if row[1] else "Imported", "created_at": str(row[2]) if row[2] else datetime.utcnow().isoformat() + "Z"})

    if not projects:
        raise ValueError("No projects in workbook")

    ws_tasks = get_sheet("Tasks")
    rows_tasks = list(ws_tasks.iter_rows(min_row=2, values_only=True))
    tasks = []
    for row in rows_tasks:
        if row and row[0]:
            scheduling_mode = "fixed"
            if len(row) >= 19 and row[13] and str(row[13]).strip().lower() in ("fixed", "auto"):
                scheduling_mode = str(row[13]).strip().lower()
            tasks.append({
                "uid": str(row[0]).strip(),
                "project_uid": str(row[1]).strip() if row[1] else projects[0]["uid"],
                "parent_task_uid": str(row[2]).strip() if row[2] else None,
                "name": str(row[3]) if row[3] else "Task",
                "description": str(row[4]) if row[4] else "",
                "accountable_person": str(row[5]) if row[5] else "",
                "responsible_party": str(row[6]) if row[6] else "",
                "start_date": str(row[7]) if row[7] else None,
                "end_date": str(row[8]) if row[8] else None,
                "is_milestone": _as_bool(row[9]) if len(row) > 14 else False,
                "status": str(row[10]) if len(row) > 14 and row[10] else (str(row[9]) if row[9] else "not_started"),
                "progress": int(row[11]) if len(row) > 14 and row[11] is not None else (int(row[10]) if row[10] is not None else 0),
                "sort_order": int(row[12]) if len(row) > 14 and row[12] is not None else (int(row[11]) if row[11] is not None else 0),
                "scheduling_mode": scheduling_mode,
                "is_deleted": _as_bool(row[14]) if len(row) > 18 else (_as_bool(row[13]) if len(row) > 17 else False),
                "deleted_at": str(row[15]) if len(row) > 18 and row[15] else (str(row[14]) if len(row) > 17 and row[14] else None),
                "deleted_by": str(row[16]) if len(row) > 18 and row[16] else (str(row[15]) if len(row) > 17 and row[15] else None),
                "created_at": str(row[17]) if len(row) > 18 and row[17] else (str(row[16]) if len(row) > 17 and row[16] else (str(row[13]) if len(row) > 13 and row[13] else datetime.utcnow().isoformat() + "Z")),
                "updated_at": str(row[18]) if len(row) > 18 and row[18] else (str(row[17]) if len(row) > 17 and row[17] else (str(row[14]) if len(row) > 14 and row[14] else datetime.utcnow().isoformat() + "Z")),
            })

    ws_dep = get_sheet("Dependencies")
    rows_dep = list(ws_dep.iter_rows(min_row=2, values_only=True))
    deps = []
    for row in rows_dep:
        if row and row[0] and row[1] and row[2] and row[3] and row[4]:
            deps.append({
                "uid": str(row[0]).strip(),
                "project_uid": str(row[1]).strip(),
                "predecessor_task_uid": str(row[2]).strip(),
                "successor_task_uid": str(row[3]).strip(),
                "dependency_type": str(row[4]).strip(),
                "created_at": str(row[5]) if len(row) > 5 and row[5] else datetime.utcnow().isoformat() + "Z",
            })

    ws_rag = get_sheet("RAG Status History")
    rows_rag = list(ws_rag.iter_rows(min_row=2, values_only=True))
    rag = []
    for row in rows_rag:
        if row and row[0] and row[1] and row[2]:
            rag.append({
                "uid": str(row[0]).strip(),
                "task_uid": str(row[1]).strip(),
                "status": str(row[2]).strip(),
                "rationale": str(row[3]) if len(row) > 3 and row[3] else "",
                "path_to_green": str(row[4]) if len(row) > 5 and row[4] else "",
                "created_at": (
                    str(row[5]) if len(row) > 5 and row[5]
                    else str(row[4]) if len(row) > 4 and row[4]
                    else datetime.utcnow().isoformat() + "Z"
                ),
            })

    ws_com = get_sheet("Comments")
    rows_com = list(ws_com.iter_rows(min_row=2, values_only=True))
    comments = []
    for row in rows_com:
        if row and row[0] and row[1]:
            comments.append({
                "uid": str(row[0]).strip(),
                "task_uid": str(row[1]).strip(),
                "author": str(row[2]) if len(row) > 2 and row[2] else "",
                "comment_text": str(row[3]) if len(row) > 3 and row[3] else "",
                "created_at": str(row[4]) if len(row) > 4 and row[4] else datetime.utcnow().isoformat() + "Z",
            })

    ws_risk = get_sheet("Risks")
    rows_risk = list(ws_risk.iter_rows(min_row=2, values_only=True))
    risks = []
    for row in rows_risk:
        if row and row[0] and row[1] and row[2]:
            risks.append({
                "uid": str(row[0]).strip(),
                "task_uid": str(row[1]).strip(),
                "title": str(row[2]) if row[2] else "Risk",
                "description": str(row[3]) if len(row) > 3 and row[3] else "",
                "severity": str(row[4]) if len(row) > 4 and row[4] else "medium",
                "status": str(row[5]) if len(row) > 5 and row[5] else "open",
                "owner": str(row[6]) if len(row) > 6 and row[6] else "",
                "mitigation_plan": str(row[7]) if len(row) > 7 and row[7] else "",
                "created_at": str(row[8]) if len(row) > 8 and row[8] else datetime.utcnow().isoformat() + "Z",
                "updated_at": str(row[9]) if len(row) > 9 and row[9] else datetime.utcnow().isoformat() + "Z",
            })

    ws_audit = get_optional_sheet("Audit Log")
    audit_events = []
    if ws_audit is not None:
        rows_audit = list(ws_audit.iter_rows(min_row=2, values_only=True))
        for row in rows_audit:
            if row and row[0]:
                audit_events.append({
                    "uid": str(row[0]).strip(),
                    "actor_employee_id": str(row[1]) if len(row) > 1 and row[1] else "SYSTEM",
                    "action_type": str(row[2]) if len(row) > 2 and row[2] else "imported",
                    "entity_type": str(row[3]) if len(row) > 3 and row[3] else "task",
                    "entity_uid": str(row[4]) if len(row) > 4 and row[4] else None,
                    "task_uid": str(row[5]) if len(row) > 5 and row[5] else None,
                    "task_name": str(row[6]) if len(row) > 6 and row[6] else None,
                    "prior_value": str(row[7]) if len(row) > 7 and row[7] else None,
                    "new_value": str(row[8]) if len(row) > 8 and row[8] else None,
                    "metadata": str(row[9]) if len(row) > 9 and row[9] else None,
                    "created_at": str(row[10]) if len(row) > 10 and row[10] else datetime.utcnow().isoformat() + "Z",
                })

    ws_lock = get_optional_sheet("Edit Lock")
    edit_lock = None
    if ws_lock is not None:
        rows_lock = list(ws_lock.iter_rows(min_row=2, values_only=True))
        for row in rows_lock:
            if row and row[0]:
                edit_lock = {
                    "lock_name": str(row[0]).strip(),
                    "employee_id": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                    "locked_at": str(row[2]) if len(row) > 2 and row[2] else datetime.utcnow().isoformat() + "Z",
                    "updated_at": str(row[3]) if len(row) > 3 and row[3] else datetime.utcnow().isoformat() + "Z",
                }
                break

    wb.close()

    # Insert in order: projects, tasks, dependencies, RAG, comments, risks
    with get_conn() as conn:
        conn.execute("DELETE FROM audit_events")
        project_uids = [p["uid"] for p in projects]
        if project_uids:
            project_placeholders = ",".join(["?"] * len(project_uids))
            existing_task_uids = [
                row["uid"]
                for row in conn.execute(
                    f"SELECT uid FROM tasks WHERE project_uid IN ({project_placeholders})",
                    tuple(project_uids),
                ).fetchall()
            ]
            imported_task_uids = [t["uid"] for t in tasks]
            all_task_uids = sorted(set(existing_task_uids + imported_task_uids))
            conn.execute(
                f"DELETE FROM dependencies WHERE project_uid IN ({project_placeholders})",
                tuple(project_uids),
            )
            if all_task_uids:
                task_placeholders = ",".join(["?"] * len(all_task_uids))
                conn.execute(f"DELETE FROM rag_statuses WHERE task_uid IN ({task_placeholders})", tuple(all_task_uids))
                conn.execute(f"DELETE FROM comments WHERE task_uid IN ({task_placeholders})", tuple(all_task_uids))
                conn.execute(f"DELETE FROM risks WHERE task_uid IN ({task_placeholders})", tuple(all_task_uids))
                conn.execute(
                    f"""DELETE FROM dependencies
                        WHERE predecessor_task_uid IN ({task_placeholders})
                           OR successor_task_uid IN ({task_placeholders})""",
                    tuple(all_task_uids) + tuple(all_task_uids),
                )
            conn.execute(f"DELETE FROM tasks WHERE project_uid IN ({project_placeholders})", tuple(project_uids))

        conn.execute("DELETE FROM edit_lock")

        for p in projects:
            conn.execute("INSERT OR REPLACE INTO projects (uid, name, created_at) VALUES (?, ?, ?)", (p["uid"], p["name"], p["created_at"]))
        for t in tasks:
            conn.execute(
                """INSERT OR REPLACE INTO tasks (uid, project_uid, parent_task_uid, name, description, accountable_person, responsible_party, start_date, end_date, is_milestone, status, progress, sort_order, scheduling_mode, is_deleted, deleted_at, deleted_by, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (t["uid"], t["project_uid"], t["parent_task_uid"], t["name"], t["description"], t["accountable_person"], t["responsible_party"], t["start_date"], t["end_date"], int(t.get("is_milestone", False)), t["status"], t["progress"], t["sort_order"], t.get("scheduling_mode", "fixed"), int(t.get("is_deleted", False)), t.get("deleted_at"), t.get("deleted_by"), t["created_at"], t["updated_at"]),
            )
        for d in deps:
            conn.execute(
                "INSERT OR REPLACE INTO dependencies (uid, project_uid, predecessor_task_uid, successor_task_uid, dependency_type, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (d["uid"], d["project_uid"], d["predecessor_task_uid"], d["successor_task_uid"], d["dependency_type"], d["created_at"]),
            )
        for r in rag:
            conn.execute(
                "INSERT OR REPLACE INTO rag_statuses (uid, task_uid, status, rationale, path_to_green, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (r["uid"], r["task_uid"], r["status"], r["rationale"], r.get("path_to_green", ""), r["created_at"])
            )
        for c in comments:
            conn.execute("INSERT OR REPLACE INTO comments (uid, task_uid, author, comment_text, created_at) VALUES (?, ?, ?, ?, ?)", (c["uid"], c["task_uid"], c["author"], c["comment_text"], c["created_at"]))
        for r in risks:
            conn.execute(
                "INSERT OR REPLACE INTO risks (uid, task_uid, title, description, severity, status, owner, mitigation_plan, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (r["uid"], r["task_uid"], r["title"], r["description"], r["severity"], r["status"], r["owner"], r["mitigation_plan"], r["created_at"], r["updated_at"]),
            )
        for event in audit_events:
            conn.execute(
                """INSERT OR REPLACE INTO audit_events (uid, actor_employee_id, action_type, entity_type, entity_uid, task_uid, task_name, prior_value, new_value, metadata, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (event["uid"], event["actor_employee_id"], event["action_type"], event["entity_type"], event["entity_uid"], event["task_uid"], event["task_name"], event["prior_value"], event["new_value"], event["metadata"], event["created_at"]),
            )
        if edit_lock and edit_lock["employee_id"]:
            conn.execute(
                "INSERT OR REPLACE INTO edit_lock (lock_name, employee_id, locked_at, updated_at) VALUES (?, ?, ?, ?)",
                (edit_lock["lock_name"], edit_lock["employee_id"], edit_lock["locked_at"], edit_lock["updated_at"]),
            )

    return {
        "projects": len(projects),
        "tasks": len(tasks),
        "dependencies": len(deps),
        "rag": len(rag),
        "comments": len(comments),
        "risks": len(risks),
        "audit_events": len(audit_events),
        "edit_lock": 1 if edit_lock and edit_lock["employee_id"] else 0,
    }
