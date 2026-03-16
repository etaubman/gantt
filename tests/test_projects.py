"""Tests for project API."""
import io

import pytest


def test_list_projects(client):
    r = client.get("/api/projects")
    assert r.status_code == 200
    data = r.json()
    assert isinstance(data, list)
    assert len(data) == 1
    assert data[0]["uid"] == "markets-data-governance"
    assert "Markets Data Governance" in data[0]["name"]
    assert "created_at" in data[0]


def test_get_single_project(client):
    r = client.get("/api/project")
    assert r.status_code == 200
    data = r.json()
    assert data["uid"] == "markets-data-governance"
    assert data["name"] == "Markets Data Governance"
    assert "created_at" in data


def test_get_project_by_uid(client, project_uid):
    r = client.get(f"/api/projects/{project_uid}")
    assert r.status_code == 200
    data = r.json()
    assert data["uid"] == project_uid
    assert data["name"] == "Markets Data Governance"


def test_get_project_not_found(client):
    r = client.get("/api/projects/nonexistent-uid")
    assert r.status_code == 404


def test_create_project_rejected(client):
    r = client.post("/api/projects", json={"name": "Another"})
    assert r.status_code == 400
    assert "only one project" in r.json().get("detail", "").lower() or "only one project" in str(r.json())


def test_delete_default_project_rejected(client, project_uid):
    r = client.delete(f"/api/projects/{project_uid}")
    assert r.status_code == 400
    assert "cannot be deleted" in r.json().get("detail", "").lower() or "cannot" in str(r.json()).lower()


def test_delete_project_not_found(client):
    r = client.delete("/api/projects/nonexistent-uid")
    assert r.status_code == 404


def test_export_by_project_uid(client, project_uid):
    r = client.get(f"/api/projects/{project_uid}/export")
    assert r.status_code == 200
    assert len(r.content) > 500
    assert "spreadsheet" in r.headers.get("content-type", "").lower() or "xlsx" in r.headers.get("content-type", "").lower()


def test_export_report_contains_overview(client):
    """Export report has Overview and Task Report sheets."""
    r = client.get("/api/export-report")
    assert r.status_code == 200
    assert len(r.content) > 500
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    assert "Overview" in wb.sheetnames
    assert "Task Report" in wb.sheetnames
    wb.close()


def test_export_project_not_found(client):
    r = client.get("/api/projects/nonexistent-uid/export")
    assert r.status_code == 404


def test_export_report_project_not_found(client):
    r = client.get("/api/projects/bad-uid/export-report")
    assert r.status_code == 404
