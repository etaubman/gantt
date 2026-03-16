"""Tests for Excel export and import API."""
import io

import pytest
from openpyxl import Workbook


def _minimal_export_workbook():
    """Build a minimal workbook that import_xlsx accepts (required sheets)."""
    wb = Workbook()
    wb.active.title = "Metadata"
    wb.active.append(["schema_version", "exported_at", "application_version"])
    wb.active.append(["3", "2025-01-01T00:00:00Z", "1.0.0"])

    ws_proj = wb.create_sheet("Projects")
    ws_proj.append(["Project UID", "Project Name", "Created At"])
    ws_proj.append(["test-proj-1", "Test Project", "2025-01-01T00:00:00Z"])

    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks.append([
        "Task UID", "Project UID", "Parent Task UID", "Name", "Description",
        "Accountable Person", "Responsible Party", "Start Date", "End Date",
        "Is Milestone", "Status", "Progress", "Sort Order",
        "Is Deleted", "Deleted At", "Deleted By", "Created At", "Updated At",
    ])
    ws_tasks.append([
        "task-1", "test-proj-1", "", "Task 1", "", "", "",
        "2025-01-01", "2025-01-15", False, "not_started", 0, 1,
        False, "", "", "2025-01-01T00:00:00Z", "2025-01-01T00:00:00Z",
    ])

    ws_dep = wb.create_sheet("Dependencies")
    ws_dep.append(["Dependency UID", "Project UID", "Predecessor Task UID", "Successor Task UID", "Dependency Type", "Created At"])

    ws_rag = wb.create_sheet("RAG Status History")
    ws_rag.append(["RAG UID", "Task UID", "Status", "Rationale", "Path To Green", "Created At"])

    ws_com = wb.create_sheet("Comments")
    ws_com.append(["Comment UID", "Task UID", "Author", "Comment Text", "Created At"])

    ws_risk = wb.create_sheet("Risks")
    ws_risk.append([
        "Risk UID", "Task UID", "Title", "Description", "Severity", "Status",
        "Owner", "Mitigation Plan", "Created At", "Updated At",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def test_export_returns_xlsx(client, project_uid):
    r = client.get("/api/export")
    assert r.status_code == 200
    assert "application/vnd.openxmlformats" in r.headers.get("content-type", "")
    assert "attachment" in r.headers.get("content-disposition", "").lower() or "export" in r.headers.get("content-disposition", "").lower()
    assert len(r.content) > 500


def test_export_report_returns_xlsx(client, project_uid):
    r = client.get("/api/export-report")
    assert r.status_code == 200
    assert "spreadsheet" in r.headers.get("content-type", "").lower() or "xlsx" in r.headers.get("content-type", "").lower()
    assert len(r.content) > 500


def test_import_accepts_valid_xlsx(client):
    content = _minimal_export_workbook()
    r = client.post(
        "/api/import",
        files={"file": ("import.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()
    assert "projects" in data
    assert "tasks" in data
    assert data["projects"] >= 1
    assert data["tasks"] >= 1


def test_import_rejects_non_excel(client):
    r = client.post(
        "/api/import",
        files={"file": ("data.txt", b"not excel", "text/plain")},
    )
    assert r.status_code == 400
    assert "excel" in r.json().get("detail", "").lower() or "xlsx" in r.json().get("detail", "").lower()


def test_import_missing_sheet_raises(client):
    wb = Workbook()
    wb.active.title = "Projects"
    wb.active.append(["UID", "Name", "Created"])
    wb.active.append(["p1", "P1", "2025-01-01"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post(
        "/api/import",
        files={"file": ("bad.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400
    # Missing sheet (e.g. Tasks) causes ValueError in import_xlsx
    assert "detail" in r.json()


def test_import_returns_counts(client):
    content = _minimal_export_workbook()
    r = client.post(
        "/api/import",
        files={"file": ("import.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()
    assert "projects" in data and data["projects"] >= 1
    assert "tasks" in data and data["tasks"] >= 1
    assert "dependencies" in data
    assert "rag" in data
    assert "comments" in data
    assert "risks" in data


def test_import_empty_filename(client):
    r = client.post(
        "/api/import",
        files={"file": ("", _minimal_export_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    # Empty filename may yield 200, 400 (bad request), or 422 (validation)
    assert r.status_code in (200, 400, 422)


def test_export_content_has_metadata_sheet(client):
    r = client.get("/api/export")
    assert r.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    assert "Metadata" in wb.sheetnames
    assert "Projects" in wb.sheetnames
    assert "Tasks" in wb.sheetnames
    wb.close()
